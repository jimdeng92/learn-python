import datetime

import os

import openpyxl

current_dir = os.path.dirname(os.path.abspath(__file__))

file_path = os.path.join(current_dir, '阿里巴巴2020年股票数据.xlsx')

# 加载一个工作簿 ---> Workbook
wb = openpyxl.load_workbook(file_path)
# 获取工作表的名字
print(wb.sheetnames)
# 获取工作表 ---> Worksheet
sheet = wb.worksheets[0]
# 获得单元格的范围
print(sheet.dimensions)
# 获得行数和列数
print(sheet.max_row, sheet.max_column)

# 获取指定单元格的值
print(sheet.cell(3, 3).value)
print(sheet['C3'].value)
print(sheet['G255'].value)
print(sheet['A2:C5'])

for row_ch in range(2, sheet.max_row + 1):
    for col_ch in 'ABCDEFG':
        value = sheet[f'{col_ch}{row_ch}'].value
        if type(value) == datetime.datetime:
            print(value.strftime('%Y年%m月%d日'), end='\t')
        elif type(value) == int:
            print(f'{value:<10d}', end='\t')
        elif type(value) == float:
            print(f'{value:.4f}', end='\t')
        else:
            print(f'{value}', end='\t')
    print()