import random

import openpyxl

wb = openpyxl.Workbook()

sheet = wb.active
sheet.title = '期末成绩'

titles = ('姓名', '语文', '数学', '英语')

for col_index, title in enumerate(titles):
    sheet.cell(row=1, column=col_index + 1).value = title

names = ('关羽', '张飞', '赵云', '马超', '黄忠')

for row_index, name in enumerate(names):
    sheet.cell(row=row_index + 2, column=1).value = name
    for col_index in range(2, 5):
        sheet.cell(row=row_index + 2, column=col_index).value = random.randrange(1, 100)


wb.save('考试成绩表.xlsx')